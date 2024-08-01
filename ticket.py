import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl


class TestSiriusSystem(unittest.TestCase):
    def setUp(self):
        self.driver = webdriver.Chrome()
        self.driver.implicitly_wait(10)

    def tearDown(self):
        self.driver.quit()

    def login(self, url, username, password):
        driver = self.driver
        driver.get(url)
        # Assuming there are input fields for username and password
        driver.find_element(By.ID, 'username').send_keys(username)
        driver.find_element(By.ID, 'password').send_keys(password)
        driver.find_element(By.ID, 'loginButton').click()

    def test_submit_application_and_approve(self):
        driver = self.driver
        driver.get("https://my-dev.internal.sirius.online/activity-page/sirius:test-tickets:v0")

        # Click on "Подать заявку"
        driver.find_element(By.XPATH, '//button[text()="Подать заявку"]').click()

        # Assuming login is required
        self.login("https://login-page-url", "your-username", "your-password")

        # Fill out the form
        driver.find_element(By.NAME, 'field1').send_keys('value1')
        driver.find_element(By.NAME, 'field2').send_keys('value2')
        driver.find_element(By.XPATH, '//button[text()="Подать заявку"]').click()

        # Refresh the page and check status
        driver.refresh()
        status = driver.find_element(By.ID, 'status').text
        self.assertEqual(status, "Рассмотрение заявки")

        # Approve the application
        driver.find_element(By.XPATH, '//artifact[text()="Одобрить"]').click()

        # Check status again
        driver.refresh()
        status = driver.find_element(By.ID, 'status').text
        self.assertEqual(status, "Заявка одобрена. Сбор документов")

        # Fill next form "Проживание"
        driver.find_element(By.XPATH, '//button[text()="Заполнить"]').click()
        driver.find_element(By.NAME, 'field1').send_keys('value1')
        driver.find_element(By.NAME, 'field2').send_keys('value2')
        driver.find_element(By.XPATH, '//button[text()="Зарегистрироваться"]').click()

        # Approve again
        driver.find_element(By.XPATH, '//artifact[text()="Одобрить"]').click()

        # Fill next form again
        driver.find_element(By.XPATH, '//button[text()="Заполнить"]').click()
        driver.find_element(By.NAME, 'field1').send_keys('value1')
        driver.find_element(By.NAME, 'field2').send_keys('value2')
        driver.find_element(By.XPATH, '//button[text()="Зарегистрироваться"]').click()

    def test_upload_tickets_via_xlsx(self):
        driver = self.driver
        driver.get("https://office-dev.internal.sirius.online/events/table/pager_size=10:pager_page=0")

        # Assuming login is required
        self.login("https://login-page-url", "your-username", "your-password")

        driver.get("https://office-dev.internal.sirius.online/events/1285/table/pager_size=10:pager_page=0")
        driver.find_element(By.XPATH, '//button[text()="⋮"]').click()
        driver.find_element(By.XPATH, '//span[text()="Действия"]').click()
        driver.find_element(By.XPATH, '//span[text()="Загрузить информацию по билетам"]').click()
        driver.find_element(By.XPATH, '//span[text()="Скачать заявки"]').click()

        # Assuming the file is downloaded to a specific directory
        xlsx_path = '/path/to/downloaded/file.xlsx'
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        # Fill in ticket information
        ws['A2'] = 'Рейс'
        ws['B2'] = 'Вагон'
        # Continue filling out the necessary fields...

        ws['N2'] = 'Нет'  # Билет куплен

        # Save the file
        wb.save(xlsx_path)

        # Upload the file
        driver.get("https://office-dev.internal.sirius.online/events/1285/table/pager_size=10:pager_page=0")
        driver.find_element(By.XPATH, '//button[text()="⋮"]').click()
        driver.find_element(By.XPATH, '//span[text()="Действия"]').click()
        driver.find_element(By.XPATH, '//span[text()="Загрузить информацию по билетам"]').click()
        driver.find_element(By.XPATH, '//span[text()="Загрузить и выполнить"]').click()

        upload_element = driver.find_element(By.XPATH, '//input[@type="file"]')
        upload_element.send_keys(xlsx_path)

        # Verify the data in the table
        driver.get(
            "https://office-dev.internal.sirius.online/registrations/1273/actions/5521/upload-tickets-info/pager_size=10:pager_page=0")
        columns = driver.find_elements(By.XPATH, '//table/thead/tr/th')
        column_texts = [col.text for col in columns]
        expected_columns = ['Рейс', 'Вагон', 'Дата', 'Время', 'Приезд. Сегмент', 'ФИО сопровождающего',
                            'Точка прибытия', 'Рейс', 'Вагон', 'Дата', 'Время', 'Отъезд.Сегмент', 'ФИО сопровождающего',
                            'Точка отбытия', 'Билет куплен', 'Комментарий']
        self.assertEqual(column_texts, expected_columns)

        # Verify the data loaded correctly
        rows = driver.find_elements(By.XPATH, '//table/tbody/tr')
        self.assertGreater(len(rows), 0)

    # Similar functions for the remaining scenarios...


if __name__ == "__main__":
    unittest.main()